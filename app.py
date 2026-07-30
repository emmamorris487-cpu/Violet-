import difflib
import re
from pathlib import Path

import openpyxl
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

app = FastAPI()

# Enable CORS for front-end requests
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- LOAD Q&A DATA FROM EXCEL ----------
# ACB_FACTS.xlsx must sit in the same folder as this file (app.py).
# It needs two columns with headers "User_Questions" and "Bot_Response".
FACTS_PATH = Path(__file__).parent / "ACB_FACTS.xlsx"

# Common filler words we ignore when comparing questions, so that
# "How do I apply for a smart card?" and "apply smart card" still match well.
STOPWORDS = {
    "a", "an", "the", "is", "are", "do", "does", "did", "i", "you", "your",
    "my", "me", "to", "for", "of", "in", "on", "at", "and", "or", "how",
    "what", "where", "when", "why", "can", "could", "would", "should",
    "please", "acb", "get", "have", "has", "with", "about", "it", "this",
    "that",
}


def normalize(text: str) -> str:
    """Lowercase and strip punctuation so text compares cleanly."""
    text = text.lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def keywords(text: str) -> set:
    return {w for w in normalize(text).split() if w and w not in STOPWORDS}


def load_faq(path: Path):
    """Read the Question/Answer pairs from the spreadsheet into memory."""
    faq = []
    if not path.exists():
        print(f"WARNING: {path} not found. Chatbot will have no answers.")
        return faq

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active  # first sheet

    # Find the header row and the two relevant columns by name,
    # so column order in the sheet doesn't matter.
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    try:
        q_col = header.index("User_Questions")
        a_col = header.index("Bot_Response")
    except ValueError:
        # Fallback: assume first two columns are question/answer
        q_col, a_col = 0, 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or q_col >= len(row) or a_col >= len(row):
            continue
        question, answer = row[q_col], row[a_col]
        if not question or not answer:
            continue
        question, answer = str(question).strip(), str(answer).strip()
        faq.append(
            {
                "question": question,
                "answer": answer,
                "normalized": normalize(question),
                "keywords": keywords(question),
            }
        )
    return faq


FAQ_DATA = load_faq(FACTS_PATH)

FALLBACK_RESPONSE = (
    "I'm not totally sure about that one. Could you rephrase your question, "
    "or contact ACB customer service at 1-268-481-4200 (headquarters) or "
    "1-268-480-1177 (Customer Service), or email customerservice@acbcaribbean.com."
)

# Minimum similarity score (0-1) required before we trust a match.
MATCH_THRESHOLD = 0.4


def find_best_answer(user_message: str) -> str:
    if not FAQ_DATA:
        return FALLBACK_RESPONSE

    norm_msg = normalize(user_message)
    msg_keywords = keywords(user_message)

    best_score = 0.0
    best_answer = None

    for entry in FAQ_DATA:
        # Text-similarity score (catches typos / close phrasing)
        text_score = difflib.SequenceMatcher(
            None, norm_msg, entry["normalized"]
        ).ratio()

        # Keyword-overlap score (catches reworded / shortened questions)
        if msg_keywords and entry["keywords"]:
            overlap = msg_keywords & entry["keywords"]
            keyword_score = len(overlap) / max(
                len(msg_keywords), len(entry["keywords"])
            )
        else:
            keyword_score = 0.0

        # Combine both signals; keyword overlap weighted a bit higher
        # since these are short FAQ-style questions.
        score = (0.4 * text_score) + (0.6 * keyword_score)

        if score > best_score:
            best_score = score
            best_answer = entry["answer"]

    if best_answer and best_score >= MATCH_THRESHOLD:
        return best_answer

    return FALLBACK_RESPONSE


# ---------- ROOT: serve the chat page ----------
# FastAPI will read index.html from the SAME folder as app.py and send it
# to the browser when someone visits your Render URL.
@app.get("/", include_in_schema=False)
def home():
    return FileResponse("index.html")


class ChatRequest(BaseModel):
    message: str


@app.post("/api/chat")
async def chat_endpoint(request: ChatRequest):
    user_message = request.message
    bot_reply = find_best_answer(user_message)
    return {"reply": bot_reply}


# Optional: reload the spreadsheet without restarting the server,
# e.g. after you edit ACB_FACTS.xlsx and re-upload it.
@app.post("/api/reload-facts", include_in_schema=False)
async def reload_facts():
    global FAQ_DATA
    FAQ_DATA = load_faq(FACTS_PATH)
    return {"status": "reloaded", "count": len(FAQ_DATA)}
